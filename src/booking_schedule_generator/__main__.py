import os
import sys

from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QVBoxLayout,
    QFormLayout,
    QLineEdit,
    QPushButton,
    QFileDialog,
    QLabel,
)
from PyQt5.QtCore import QStandardPaths, Qt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
import locale
import calendar
from block import Block
from data.data import appartments, seasons
import os


class Dashboard(QMainWindow):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Buchungsplan Generator")
        self.setGeometry(100, 100, 400, 200)

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)

        self.layout = QVBoxLayout(self.central_widget)

        self.header_label = QLabel("<h2>Buchungsplan Generator</h2>")
        self.header_label.setAlignment(Qt.AlignCenter)
        self.layout.addWidget(self.header_label)

        self.form_layout = QFormLayout()
        self.year_input = QLineEdit()
        self.year_input.setPlaceholderText("z.B. 2025")

        desktop_path = QStandardPaths.writableLocation(QStandardPaths.DesktopLocation)
        self.save_location_label = QLabel(desktop_path)

        self.form_layout.addRow("Jahr:", self.year_input)
        self.form_layout.addRow("Speichern unter:", self.save_location_label)
        self.layout.addLayout(self.form_layout)

        self.save_location_button = QPushButton("Speicherort wählen")
        self.save_location_button.clicked.connect(self.browse_location)
        self.layout.addWidget(self.save_location_button)

        self.generate_button = QPushButton("Buchungsplan generieren")
        self.generate_button.clicked.connect(self.generate_file)
        self.layout.addWidget(self.generate_button)

        self.status_label = QLabel()
        self.layout.addWidget(self.status_label)

    def browse_location(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Folder")
        if folder:
            self.save_location_label.setText(folder)

    def generate_file(self):
        try:
            year = int(self.year_input.text())
            save_location = self.save_location_label.text()
            if not save_location or save_location == "No location selected":
                self.status_label.setText("Please select a save location.")
                return

            locale.setlocale(locale.LC_ALL, "de_AT")
            property_name = "Landhaus Zoppoth"
            primary_color = "ff8d42"
            secondary_color = "4472c4"

            wb = Workbook()
            for i in range(0, 12):
                if i == 0:
                    ws = wb.active
                    ws.title = calendar.month_name[i + 1]
                else:
                    ws = wb.create_sheet(calendar.month_name[i + 1])

                ws.merge_cells("A1:C1")
                ws["A1"].value = property_name
                ws["A1"].font = Font(
                    bold=True, italic=True, color=primary_color, size=18
                )
                ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
                ws["A1"].fill = PatternFill(
                    start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
                )

                ws.merge_cells("A2:C2")
                ws["A2"].value = f"{calendar.month_name[i + 1]} - {year}"
                ws["A2"].font = Font(
                    bold=True, italic=True, color=secondary_color, size=12
                )
                ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
                ws["A2"].fill = PatternFill(
                    start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
                )

                ws.merge_cells("A3:C3")
                ws["A3"].fill = PatternFill(
                    start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
                )

                block = Block(
                    month=i + 1,
                    year=year,
                    worksheet=ws,
                    primary_color=primary_color,
                    secondary_color=secondary_color,
                    seasons=seasons,
                )

                block.date_block(start_col=4, start_row=2)
                block.blank_block(start_col=4, start_row=4, end_row=8)
                block.price_block(
                    start_col=1, start_row=4, appartment=appartments["wg_1"]
                )

            file_path = f"{save_location}/Wohnungsplan-{year}.xlsx"
            wb.save(file_path)
            self.status_label.setText(f"File saved to {file_path}")

        except ValueError:
            self.status_label.setText("Invalid year. Please enter a numeric value.")


def main():
    app = QApplication(sys.argv)
    dashboard = Dashboard()
    dashboard.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
