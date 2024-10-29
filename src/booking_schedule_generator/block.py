from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
import calendar


class Block:
    def __init__(
        self,
        month,
        year,
        worksheet,
        primary_color,
        secondary_color,
        seasons: dict,
        lang: str = "DE",
        separate_week_at: int = 5,
    ):
        self.month = month
        self.year = year
        self.days = self.get_num_days(year, month)
        self.worksheet = worksheet
        self.primary_color = primary_color
        self.secondary_color = secondary_color
        self.medium_side = Side(style="medium")
        self.dotted_side = Side(style="dotted")
        self.seasons = seasons
        self.separate_week_at = separate_week_at

        weekdays_by_lang = {
            "DE": ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"],
            "EN": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"],
        }

        self.weekdays = weekdays_by_lang.get(lang)
        self.current_seasons: list = self.get_season()
        self.season_change_day = self.check_for_season_change()

    def get_num_days(self, year, month):
        """Calculate and return the number of days in the given month and year."""
        _, num_days = calendar.monthrange(year, month)
        return num_days

    def check_for_season_change(self) -> int:
        """Checks if month contains season change"""
        for season in self.seasons.values():
            for date in season.values():
                if date.month == self.month:
                    return date.day
        else:
            return 0

    def get_season(self):
        month_to_seasons = {
            1: ["christmas", "winter_past"],
            2: ["winter_past"],
            3: ["winter_past"],
            4: ["spring"],
            5: ["spring"],
            6: ["spring"],
            7: ["summer"],
            8: ["summer"],
            9: ["summer", "fall"],
            10: ["fall"],
            11: ["winter_pre"],
            12: ["winter_pre", "christmas"],
        }

        return month_to_seasons.get(self.month)

    def date_block(self, start_col, start_row):
        """Generates the date block including week day names for the given month ans year."""

        for i in range(1, self.days + 1):
            weekday_index = calendar.weekday(self.year, self.month, i)
            weekday_name = self.weekdays[weekday_index]

            day_cell = self.worksheet.cell(row=start_row, column=start_col + i - 1)
            date_cell = self.worksheet.cell(row=start_row + 1, column=start_col + i - 1)
            day_cell.value = weekday_name
            date_cell.value = i
            day_cell.alignment = Alignment(horizontal="center")
            date_cell.alignment = Alignment(horizontal="center")

            day_cell.fill = PatternFill(
                start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
            )
            date_cell.fill = PatternFill(
                start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
            )

            if i == 1:
                day_cell.border = Border(
                    left=self.medium_side,
                    top=self.medium_side,
                    right=self.dotted_side,
                    bottom=None,
                )
                date_cell.border = Border(
                    left=self.medium_side,
                    top=None,
                    right=self.dotted_side,
                    bottom=self.medium_side,
                )
            elif i == self.days:
                day_cell.border = Border(
                    left=self.dotted_side,
                    top=self.medium_side,
                    right=self.medium_side,
                    bottom=None,
                )
                date_cell.border = Border(
                    left=self.dotted_side,
                    top=None,
                    right=self.medium_side,
                    bottom=self.medium_side,
                )
            else:
                if weekday_name == self.weekdays[self.separate_week_at]:
                    day_cell.border = Border(
                        left=self.dotted_side,
                        top=self.medium_side,
                        right=self.medium_side,
                        bottom=None,
                    )
                    date_cell.border = Border(
                        left=self.dotted_side,
                        top=None,
                        right=self.medium_side,
                        bottom=self.medium_side,
                    )
                else:
                    day_cell.border = Border(
                        left=self.dotted_side,
                        right=self.dotted_side,
                        top=self.medium_side,
                        bottom=None,
                    )
                    date_cell.border = Border(
                        left=self.dotted_side,
                        right=self.dotted_side,
                        top=None,
                        bottom=self.medium_side,
                    )

            if i == self.season_change_day:
                day_cell.fill = PatternFill(
                    start_color=self.primary_color,
                    end_color=self.primary_color,
                    fill_type="solid",
                )
                date_cell.fill = PatternFill(
                    start_color=self.primary_color,
                    end_color=self.primary_color,
                    fill_type="solid",
                )

    def blank_block(self, start_col, start_row, end_row):
        """Generates the blank space between the date blocks used for guest schedules."""
        for i in range(1, self.days + 1):
            weekday_index = calendar.weekday(self.year, self.month, i)
            weekday_name = self.weekdays[weekday_index]
            for r in range(start_row, end_row):
                cell = self.worksheet.cell(row=r, column=start_col + i - 1)
                cell.value = None
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill(
                    start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
                )
                if i == 1:
                    cell.border = Border(left=self.medium_side)
                elif i == self.days:
                    cell.border = Border(right=self.medium_side)
                else:
                    if (
                        weekday_name == self.weekdays[self.separate_week_at]
                    ):  # TODO: implement this more efficiently
                        cell.border = Border(right=self.medium_side)

    def price_block(self, start_col, start_row, appartment):
        if len(self.current_seasons) == 1:
            self.generate_price_block(
                start_col, start_row, appartment, self.current_seasons[0]
            )
            self.generate_price_block(
                start_col + 3 + self.days,
                start_row,
                appartment,
                self.current_seasons[0],
            )
        if len(self.current_seasons) == 2:
            self.generate_price_block(
                start_col, start_row, appartment, self.current_seasons[0]
            )
            self.generate_price_block(
                start_col + 3 + self.days,
                start_row,
                appartment,
                self.current_seasons[1],
            )

    def generate_price_block(self, start_col, start_row, appartment, season):
        white_fill = PatternFill(
            start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
        )
        font = Font(italic=True, color=self.primary_color, size=12)
        bold_font = Font(bold=True, italic=True, color=self.primary_color, size=12)
        alignment = Alignment(horizontal="center", vertical="center")

        if season:
            vertical_heigth = len(appartment["prices"][season])

            for i, (persons, price) in enumerate(appartment["prices"][season].items()):
                person_cell = self.worksheet.cell(
                    row=start_row + i, column=start_col + 1
                )
                person_cell.value = persons
                person_cell.fill = white_fill
                person_cell.font = font
                person_cell.alignment = alignment
                person_cell.border = Border(
                    left=self.dotted_side,
                    right=self.dotted_side,
                    bottom=self.dotted_side,
                )

                price_cell = self.worksheet.cell(
                    row=start_row + i, column=start_col + 2
                )
                price_cell.value = f"{price}€"
                price_cell.fill = white_fill
                price_cell.font = font
                price_cell.alignment = alignment
                price_cell.border = Border(
                    left=self.dotted_side,
                    right=self.medium_side,
                    bottom=self.dotted_side,
                )

                if i == 0:
                    person_cell.border = Border(
                        left=self.dotted_side,
                        right=self.dotted_side,
                        top=self.medium_side,
                        bottom=self.dotted_side,
                    )
                    price_cell.border = Border(
                        left=self.dotted_side,
                        right=self.medium_side,
                        top=self.medium_side,
                        bottom=self.dotted_side,
                    )

            if vertical_heigth < 3:
                compensation_cell = self.worksheet.cell(
                    row=start_row + 2, column=start_col + 2
                )
                compensation_cell.fill = white_fill
                compensation_cell.border = Border(
                    left=self.dotted_side,
                    right=self.medium_side,
                    bottom=self.dotted_side,
                )

        cleaning_price_description_cell = self.worksheet.cell(
            row=start_row + 3, column=start_col + 1
        )
        cleaning_price_description_cell.value = f"ER:"
        cleaning_price_description_cell.fill = white_fill
        cleaning_price_description_cell.font = font
        cleaning_price_description_cell.alignment = alignment
        cleaning_price_description_cell.border = Border(
            left=self.dotted_side,
            right=self.dotted_side,
            top=self.dotted_side,
            bottom=self.medium_side,
        )

        cleaning_price_cell = self.worksheet.cell(
            row=start_row + 3, column=start_col + 2
        )
        cleaning_price_cell.value = f"{appartment["cleaning_price"]}€"
        cleaning_price_cell.fill = white_fill
        cleaning_price_cell.font = font
        cleaning_price_cell.alignment = alignment
        cleaning_price_cell.border = Border(
            left=self.dotted_side,
            right=self.medium_side,
            top=self.dotted_side,
            bottom=self.medium_side,
        )

        name_cell = self.worksheet.cell(row=start_row, column=start_col)
        name_cell.value = appartment["name_full"]
        name_cell.fill = white_fill
        name_cell.font = bold_font
        name_cell.alignment = alignment
        name_cell.border = Border(
            left=self.medium_side, right=self.medium_side, top=self.medium_side
        )

        type_cell = self.worksheet.cell(row=start_row + 1, column=start_col)
        type_cell.value = appartment["type"]
        type_cell.fill = white_fill
        type_cell.font = bold_font
        type_cell.alignment = alignment
        type_cell.border = Border(left=self.medium_side, right=self.medium_side)

        size_cell = self.worksheet.cell(row=start_row + 2, column=start_col)
        size_cell.value = appartment["size"]
        size_cell.fill = white_fill
        size_cell.font = bold_font
        size_cell.alignment = alignment
        size_cell.border = Border(left=self.medium_side, right=self.medium_side)

        end_cell = self.worksheet.cell(row=start_row + 3, column=start_col)
        end_cell.fill = white_fill
        end_cell.border = Border(
            left=self.medium_side, right=self.medium_side, bottom=self.medium_side
        )
