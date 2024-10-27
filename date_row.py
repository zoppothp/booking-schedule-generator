from openpyxl.styles import Alignment
import calendar

class DateRow:
    def __init__(self, month, year, col, row):
        self.month = month
        self.year = year
        self.col = col
        self.row = row

    def setup_dates(self, ws, start_col):
        _, num_days = calendar.monthrange(self.year, self.month)
        
        for i in range(1, num_days + 1):
            weekday_index = calendar.weekday(self.year, self.month, i)
            weekday_name = ["mo", "di", "mi", "do", "fr", "sa", "so"][weekday_index]
            ws.cell(row=self.row, column=start_col + i - 1).value = weekday_name
            ws.cell(row=self.row, column=start_col + i - 1).alignment = Alignment(horizontal="center")
            ws.cell(row=self.row + 1, column=start_col + i - 1).value = i
            ws.cell(row=self.row + 1, column=start_col + i - 1).alignment = Alignment(horizontal="center")