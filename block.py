from openpyxl.styles import Alignment, Border, Side, PatternFill
import calendar


class Block:
    def __init__(
        self,
        month,
        year,
        worksheet,
        primary_color,
        secondary_color,
        prices: dict,
        seasons: dict,
        lang: str = "DE",
        separate_week_at: int = 5,
    ):
        self.month = month
        self.year = year
        self.worksheet = worksheet
        self.primary_color = primary_color
        self.secondary_color = secondary_color
        self.medium_side = Side(style="medium")
        self.dotted_side = Side(style="dotted")
        self.prices = prices
        self.seasons = seasons
        self.separate_week_at = separate_week_at

        weekdays_by_lang = {
            "DE": ["Mo", "Di", "Mi", "Do", "Fr", "Sa", "So"],
            "EN": ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"],
        }

        self.weekdays = weekdays_by_lang.get(lang)
        self.season_change = self.check_for_season_change()

    def get_num_days(self):
        """Calculate and return the number of days in the given month and year."""
        _, num_days = calendar.monthrange(self.year, self.month)
        return num_days

    def check_for_season_change(self) -> int:
        """Checks if month contains season change"""
        for season in self.seasons.values():
            for date in season:
                if date.month == self.month:
                    return date.day
        else:
            return 0

    def date_block(self, start_col, start_row):
        """Generates the date block including week day names for the given month ans year."""
        num_days = self.get_num_days()

        for i in range(1, num_days + 1):
            weekday_index = calendar.weekday(self.year, self.month, i)
            weekday_name = self.weekdays[weekday_index]

            day_cell = self.worksheet.cell(row=start_row, column=start_col + i - 1)
            date_cell = self.worksheet.cell(row=start_row + 1, column=start_col + i - 1)
            day_cell.value = weekday_name
            date_cell.value = i
            day_cell.alignment = Alignment(horizontal="center")
            date_cell.alignment = Alignment(horizontal="center")

            day_cell.fill = PatternFill(
                start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
            )
            date_cell.fill = PatternFill(
                start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
            )

            if i == 1:
                day_cell.border = Border(
                    left=self.medium_side,
                    top=self.medium_side,
                    right=self.dotted_side,
                    bottom=None,
                )
                date_cell.border = Border(
                    left=self.medium_side,
                    top=None,
                    right=self.dotted_side,
                    bottom=self.medium_side,
                )
            elif i == num_days:
                day_cell.border = Border(
                    left=self.dotted_side,
                    top=self.medium_side,
                    right=self.medium_side,
                    bottom=None,
                )
                date_cell.border = Border(
                    left=self.dotted_side,
                    top=None,
                    right=self.medium_side,
                    bottom=self.medium_side,
                )
            else:
                if weekday_name == self.weekdays[self.separate_week_at]:
                    day_cell.border = Border(
                        left=self.dotted_side,
                        top=self.medium_side,
                        right=self.medium_side,
                        bottom=None,
                    )
                    date_cell.border = Border(
                        left=self.dotted_side,
                        top=None,
                        right=self.medium_side,
                        bottom=self.medium_side,
                    )
                else:
                    day_cell.border = Border(
                        left=self.dotted_side,
                        right=self.dotted_side,
                        top=self.medium_side,
                        bottom=None,
                    )
                    date_cell.border = Border(
                        left=self.dotted_side,
                        right=self.dotted_side,
                        top=None,
                        bottom=self.medium_side,
                    )

            if i == self.season_change:
                day_cell.fill = PatternFill(
                    start_color=self.primary_color,
                    end_color=self.primary_color,
                    fill_type="solid",
                )
                date_cell.fill = PatternFill(
                    start_color=self.primary_color,
                    end_color=self.primary_color,
                    fill_type="solid",
                )

    def blank_block(self, start_col, start_row, end_row):
        """Generates the blank space between the date blocks used for guest schedules."""
        num_days = self.get_num_days()
        for i in range(1, num_days + 1):
            weekday_index = calendar.weekday(self.year, self.month, i)
            weekday_name = self.weekdays[weekday_index]
            for r in range(start_row, end_row):
                cell = self.worksheet.cell(row=r, column=start_col + i - 1)
                cell.value = None
                cell.alignment = Alignment(horizontal="center")
                cell.fill = PatternFill(
                    start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
                )
                if i == 1:
                    cell.border = Border(left=self.medium_side)
                elif i == num_days:
                    cell.border = Border(right=self.medium_side)
                else:
                    if (
                        weekday_name == self.weekdays[self.separate_week_at]
                    ):  # TODO: implement this more efficiently
                        cell.border = Border(right=self.medium_side)
