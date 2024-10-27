from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from block import Block

wb = Workbook()
ws = wb.active
ws.title = "Jänner"

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="000000")

ws.merge_cells("B1:C1")
ws["B1"].value = "Landhaus Zoppoth"
ws["B1"].font = header_font
ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
ws["B1"].fill = header_fill

block = Block(month=1, year=2025, worksheet=ws)
block.date_row(col=4, row=2)

wb.save("booking-schedule.xlsx")
