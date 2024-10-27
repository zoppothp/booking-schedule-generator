from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from date_row import DateRow  # Import the DateRow class

wb = Workbook()
ws = wb.active
ws.title = "Jänner"

header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="000000")

ws.merge_cells("B1:C1")
ws["B1"].value = "Landhaus Zoppoth"
ws["B1"].font = header_font
ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
ws["B1"].fill = header_fill

date_row_1 = DateRow(month=1, year=2025, col=4, row=3)
date_row_1.setup_dates(ws, start_col=4)

date_row_2 = DateRow(month=1, year=2025, col=4, row=6)
date_row_2.setup_dates(ws, start_col=4)


wb.save("booking-schedule.xlsx")