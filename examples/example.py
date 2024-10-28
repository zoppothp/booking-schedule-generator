from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from block import Block
import locale
import calendar
from examples.prices import appartment_typ_a_prices, seasons

locale.setlocale(locale.LC_ALL, "de_AT")
year = 2025
property_name = "Landhaus Zoppoth"
primary_color = "ff8d42"
secondary_color = "4472c4"

wb = Workbook()

for i in range(0, 12):
    if i == 0:
        ws = wb.active
        ws.title = calendar.month_name[i + 1]
    else:
        ws = wb.create_sheet(calendar.month_name[i + 1])

    ws.merge_cells("A1:C1")
    ws["A1"].value = property_name
    ws["A1"].font = Font(bold=True, italic=True, color=primary_color, size=18)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )

    ws.merge_cells("A2:C2")
    ws["A2"].value = f"{calendar.month_name[i+1]} - {year}"
    ws["A2"].font = Font(bold=True, italic=True, color=secondary_color, size=12)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A2"].fill = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )

    ws.merge_cells("A3:C3")
    ws["A3"].fill = PatternFill(
        start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"
    )

    block = Block(
        month=i + 1,
        year=year,
        worksheet=ws,
        primary_color=primary_color,
        secondary_color=secondary_color,
        prices=appartment_typ_a_prices,
        seasons=seasons,
    )

    block.date_block(start_col=4, start_row=2)
    block.date_block(start_col=4, start_row=8)
    block.date_block(start_col=4, start_row=14)
    block.date_block(start_col=4, start_row=20)
    block.date_block(start_col=4, start_row=26)
    block.date_block(start_col=4, start_row=32)
    block.date_block(start_col=4, start_row=38)

    block.blank_block(start_col=4, start_row=4, end_row=8)
    block.blank_block(start_col=4, start_row=10, end_row=14)
    block.blank_block(start_col=4, start_row=16, end_row=20)
    block.blank_block(start_col=4, start_row=22, end_row=26)
    block.blank_block(start_col=4, start_row=28, end_row=32)
    block.blank_block(start_col=4, start_row=34, end_row=38)

wb.save("booking-schedule.xlsx")
